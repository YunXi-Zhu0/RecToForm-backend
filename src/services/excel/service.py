from pathlib import Path
from shutil import copyfile
from typing import Iterable, Optional

from src.core.config import DEFAULT_MISSING_VALUE, DEFAULT_OUTPUT_DIR
from src.services.excel.models import (
    ExcelWriteRequest,
    ExcelWriteResult,
    StandardExcelWriteRequest,
    TableExcelWriteRequest,
    StructuredInvoiceData,
)
from src.services.llm.models import StructuredExtractionResult
from src.services.template.models import ExcelFieldMapping


class ExcelWriteError(ValueError):
    pass


SOURCE_FILE_FIELD_ID = "源文件"


class ExcelService:
    def __init__(self, output_dir: Optional[Path] = None) -> None:
        self.output_dir = Path(output_dir or DEFAULT_OUTPUT_DIR)

    def write_table(self, request: TableExcelWriteRequest) -> ExcelWriteResult:
        from openpyxl import Workbook

        request.output_dir.mkdir(parents=True, exist_ok=True)
        output_path = request.output_dir / self._resolve_table_output_filename(request)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = request.sheet_name

        for column_index, header in enumerate(request.headers, start=1):
            sheet.cell(row=1, column=column_index, value=header)

        for row_index, row in enumerate(request.rows, start=2):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)

        workbook.save(output_path)
        workbook.close()
        return ExcelWriteResult(
            output_file_path=output_path,
            written_fields=list(request.headers),
            skipped_fields=[],
            missing_mappings=[],
        )

    def write_standard_fields(self, request: StandardExcelWriteRequest) -> ExcelWriteResult:
        from openpyxl import Workbook

        request.output_dir.mkdir(parents=True, exist_ok=True)
        output_path = request.output_dir / self._resolve_standard_output_filename(request)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = request.sheet_name

        headers = list(request.standard_fields)
        if request.source_file_name and SOURCE_FILE_FIELD_ID not in headers:
            headers = [SOURCE_FILE_FIELD_ID] + headers

        written_fields = []
        for column_index, field_name in enumerate(headers, start=1):
            sheet.cell(row=1, column=column_index, value=field_name)
            if field_name == SOURCE_FILE_FIELD_ID:
                cell_value = request.source_file_name
            else:
                cell_value = request.structured_data.data.get(field_name, DEFAULT_MISSING_VALUE)
            sheet.cell(
                row=2,
                column=column_index,
                value=cell_value,
            )
            written_fields.append(field_name)

        workbook.save(output_path)
        workbook.close()
        return ExcelWriteResult(
            output_file_path=output_path,
            written_fields=written_fields,
            skipped_fields=[],
            missing_mappings=[],
        )

    def write(self, request: ExcelWriteRequest) -> ExcelWriteResult:
        from openpyxl import load_workbook

        self._validate_request(request)
        request.output_dir.mkdir(parents=True, exist_ok=True)
        output_path = request.output_dir / self._resolve_output_filename(request)
        copyfile(request.excel_template_path, output_path)

        workbook = load_workbook(output_path)
        self._sync_headers(workbook=workbook, request=request)

        written_fields = []
        for field_id in request.export_field_ids:
            mapping = request.excel_mappings.get(field_id)
            if mapping is None:
                continue
            value = self._resolve_value(mapping=mapping, structured_data=request.structured_data)
            workbook[mapping.sheet_name][mapping.cell] = value
            written_fields.append(field_id)

        workbook.save(output_path)
        workbook.close()
        missing_mappings = [
            field_id for field_id in request.export_field_ids if field_id not in request.excel_mappings
        ]
        skipped_fields = [field_id for field_id in request.export_field_ids if field_id not in written_fields]
        return ExcelWriteResult(
            output_file_path=output_path,
            written_fields=written_fields,
            skipped_fields=skipped_fields,
            missing_mappings=missing_mappings,
        )

    def build_structured_invoice_data(
        self,
        result: StructuredExtractionResult,
        standard_fields: Iterable[str],
    ) -> StructuredInvoiceData:
        normalized = {
            field_id: str(result.data.get(field_id, DEFAULT_MISSING_VALUE) or DEFAULT_MISSING_VALUE)
            for field_id in standard_fields
        }
        missing_fields = [
            field_id for field_id in standard_fields if normalized[field_id] == DEFAULT_MISSING_VALUE
        ]
        return StructuredInvoiceData(
            data=normalized,
            missing_fields=missing_fields,
            extra_fields=list(result.extra_fields),
        )

    def _sync_headers(self, workbook, request: ExcelWriteRequest) -> None:
        selected = set(request.export_field_ids)
        for field_id, mapping in request.excel_mappings.items():
            header_cell = self._resolve_header_cell(mapping.cell)
            if field_id in selected:
                workbook[mapping.sheet_name][header_cell] = request.default_header_labels.get(
                    field_id,
                    field_id,
                )
                continue
            workbook[mapping.sheet_name][header_cell] = ""
            workbook[mapping.sheet_name][mapping.cell] = ""

    def _resolve_value(
        self,
        mapping: ExcelFieldMapping,
        structured_data: StructuredInvoiceData,
    ) -> str:
        if mapping.value_source == "standard":
            value = structured_data.data.get(mapping.source_key, mapping.default_value)
            return str(value or mapping.default_value or DEFAULT_MISSING_VALUE)
        return mapping.default_value or DEFAULT_MISSING_VALUE

    def _resolve_header_cell(self, data_cell: str) -> str:
        from openpyxl.utils.cell import coordinate_from_string

        column_letters, row_index = coordinate_from_string(data_cell)
        if row_index <= 1:
            return "%s1" % column_letters
        return "%s%s" % (column_letters, row_index - 1)

    def _validate_request(self, request: ExcelWriteRequest) -> None:
        from openpyxl import load_workbook

        if not request.excel_template_path.is_file():
            raise ExcelWriteError("Excel template file not found: %s" % request.excel_template_path)
        workbook = load_workbook(request.excel_template_path, read_only=True)
        sheet_names = set(workbook.sheetnames)
        workbook.close()
        for field_id in request.export_field_ids:
            mapping = request.excel_mappings.get(field_id)
            if mapping is None:
                continue
            if mapping.sheet_name not in sheet_names:
                raise ExcelWriteError(
                    "Worksheet not found for field %s: %s" % (field_id, mapping.sheet_name)
                )

    def _resolve_output_filename(self, request: ExcelWriteRequest) -> str:
        if request.output_filename:
            return request.output_filename
        stem = request.excel_template_path.stem
        suffix = request.excel_template_path.suffix or ".xlsx"
        return "%s_%s%s" % (stem, request.template_id, suffix)

    def _resolve_standard_output_filename(self, request: StandardExcelWriteRequest) -> str:
        if request.output_filename:
            return request.output_filename
        return "standard_fields_export.xlsx"

    def _resolve_table_output_filename(self, request: TableExcelWriteRequest) -> str:
        if request.output_filename:
            return request.output_filename
        return "table_export.xlsx"

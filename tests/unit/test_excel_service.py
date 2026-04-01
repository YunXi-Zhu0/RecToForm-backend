from pathlib import Path

from openpyxl import load_workbook

from src.services.excel import (
    ExcelService,
    ExcelWriteRequest,
    StandardExcelWriteRequest,
    StructuredInvoiceData,
)
from src.services.template import TemplateService


def test_write_populates_template_cells(tmp_path: Path) -> None:
    template_bundle = TemplateService().get_template_bundle("finance_invoice")
    excel_service = ExcelService()
    structured_data = StructuredInvoiceData(
        data={
            "发票代码": "CODE-001",
            "发票号码": "INV-001",
            "开票日期": "",
            "购买方名称": "",
            "购买方纳税人识别号": "",
            "购买方地址电话": "",
            "购买方开户行及账号": "",
            "货物或应税劳务、服务名称": "",
            "规格型号": "",
            "单位": "",
            "数量": "",
            "单价": "",
            "金额": "",
            "税率": "",
            "税额": "",
            "合计": "199.00",
            "价税合计(大写)": "",
            "销售方名称": "",
            "销售方纳税人识别号": "",
            "销售方地址电话": "",
            "销售方开户行及账号": "",
            "收款人": "",
            "复核": "",
            "开票人": "",
            "销售方": "",
            "备注": "invoice.png",
        },
        missing_fields=[],
        extra_fields=[],
    )

    result = excel_service.write(
        ExcelWriteRequest(
            template_id=template_bundle.template_id,
            template_version=template_bundle.template_version,
            mapping_version=template_bundle.mapping_version,
            excel_template_path=template_bundle.excel_template_path,
            structured_data=structured_data,
            export_field_ids=template_bundle.recommended_field_ids,
            default_header_labels=template_bundle.default_header_labels,
            excel_mappings=template_bundle.excel_mappings,
            output_dir=tmp_path,
            output_filename="filled.xlsx",
        )
    )

    workbook = load_workbook(result.output_file_path)
    sheet = workbook["Sheet1"]
    assert sheet["A2"].value == "1"
    assert sheet["B2"].value == "INV-001"
    assert sheet["C2"].value == "CODE-001"
    assert sheet["D2"].value == "199.00"
    assert sheet["E2"].value == "invoice.png"
    assert result.missing_mappings == []


def test_asset_import_template_reads_standard_json_values(tmp_path: Path) -> None:
    template_bundle = TemplateService().get_template_bundle("asset_import")
    excel_service = ExcelService()
    structured_data = StructuredInvoiceData(
        data={
            "发票代码": "CODE-002",
            "发票号码": "INV-002",
            "开票日期": "2026-03-27",
            "购买方名称": "",
            "购买方纳税人识别号": "",
            "购买方地址电话": "",
            "购买方开户行及账号": "",
            "货物或应税劳务、服务名称": "实验耗材",
            "规格型号": "A-01",
            "单位": "件",
            "数量": "2",
            "单价": "50.00",
            "金额": "",
            "税率": "",
            "税额": "",
            "合计": "100.00",
            "价税合计(大写)": "",
            "销售方名称": "供应商A",
            "销售方纳税人识别号": "",
            "销售方地址电话": "",
            "销售方开户行及账号": "",
            "收款人": "",
            "复核": "",
            "开票人": "",
            "销售方": "",
            "备注": "asset.png",
        },
        missing_fields=[],
        extra_fields=[],
    )

    result = excel_service.write(
        ExcelWriteRequest(
            template_id=template_bundle.template_id,
            template_version=template_bundle.template_version,
            mapping_version=template_bundle.mapping_version,
            excel_template_path=template_bundle.excel_template_path,
            structured_data=structured_data,
            export_field_ids=template_bundle.recommended_field_ids,
            default_header_labels=template_bundle.default_header_labels,
            excel_mappings=template_bundle.excel_mappings,
            output_dir=tmp_path,
            output_filename="asset.xlsx",
        )
    )

    workbook = load_workbook(result.output_file_path)
    sheet = workbook["Sheet1"]
    assert sheet["B2"].value == "实验耗材"
    assert sheet["D2"].value == "A-01"
    assert sheet["H2"].value == "100.00"
    assert sheet["I2"].value == "供应商A"
    assert sheet["J2"].value == "INV-002"


def test_write_standard_fields_creates_generic_excel(tmp_path: Path) -> None:
    excel_service = ExcelService()
    structured_data = StructuredInvoiceData(
        data={
            "发票代码": "CODE-003",
            "发票号码": "INV-003",
            "备注": "generic.xlsx",
        },
        missing_fields=[],
        extra_fields=[],
    )

    result = excel_service.write_standard_fields(
        StandardExcelWriteRequest(
            structured_data=structured_data,
            standard_fields=["发票代码", "发票号码", "备注"],
            output_dir=tmp_path,
            output_filename="standard.xlsx",
            source_file_name="invoice-003.png",
        )
    )

    workbook = load_workbook(result.output_file_path)
    sheet = workbook["Sheet1"]
    assert sheet["A1"].value == "源文件"
    assert sheet["B1"].value == "发票代码"
    assert sheet["C1"].value == "发票号码"
    assert sheet["D1"].value == "备注"
    assert sheet["A2"].value == "invoice-003.png"
    assert sheet["B2"].value == "CODE-003"
    assert sheet["C2"].value == "INV-003"
    assert sheet["D2"].value == "generic.xlsx"
